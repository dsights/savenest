#!/usr/bin/env python3
"""
SaveNest Product Import Script
================================
Reads SaveNest_Product_Template.xlsx and merges/updates products.json.

Usage:
    python3 import_products.py                         # merge template → products.json
    python3 import_products.py --overwrite             # full replace (existing data lost)
    python3 import_products.py --category electricity  # only import one category
    python3 import_products.py --dry-run               # preview without writing

The script:
  1. Reads each category sheet from the Excel template.
  2. Converts each row to the Deal JSON structure.
  3. Merges into products.json (adds new, updates existing by id, skips blank rows).
  4. Writes the updated products.json back to disk.
"""

import json
import re
import sys
import argparse
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not found. Run:  pip install openpyxl")
    sys.exit(1)

# ── Paths ──────────────────────────────────────────────────────────────
SCRIPT_DIR   = Path(__file__).parent
TEMPLATE     = SCRIPT_DIR / "SaveNest_Product_Template.xlsx"
PRODUCTS_JSON= SCRIPT_DIR.parent / "products.json"

# ── Sheet → JSON category key mapping ─────────────────────────────────
SHEET_TO_CAT = {
    "⚡ Electricity": "electricity",
    "🔥 Gas":         "gas",
    "📱 Mobile":      "mobile",
    "🌐 Internet":    "internet",
    "🛡 Insurance":   "insurance",
    "☀️ Solar":       "solar",
    "💳 Credit Cards":"financial",
}

# ── Helpers ────────────────────────────────────────────────────────────

def cell_val(cell):
    """Return clean string or None."""
    v = cell.value
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None

def to_bool(val):
    """Convert cell value to bool or None."""
    if val is None:
        return None
    s = str(val).strip().upper()
    if s == "TRUE":  return True
    if s == "FALSE": return False
    return None

def to_float(val, default=0.0):
    try:
        return float(str(val).replace(",", "").replace("$", ""))
    except (TypeError, ValueError):
        return default

def clean_id(raw_id):
    """Ensure id is lowercase-underscore-slug."""
    return re.sub(r"[^a-z0-9_]", "_", str(raw_id).lower().strip())

def split_states(val):
    """'NSW,VIC,QLD' → ['NSW','VIC','QLD']"""
    if not val:
        return []
    return [s.strip().upper() for s in str(val).split(",") if s.strip()]

def parse_row(headers, row_cells):
    """
    Given a list of header names and a list of cells (same length),
    return a dict mapping header → clean value.
    """
    return {headers[i]: cell_val(row_cells[i]) for i in range(len(headers))}


def row_to_deal(data: dict) -> dict | None:
    """Convert a flat row dict to the Deal JSON structure. Returns None if row is blank."""
    raw_id = data.get("id ★") or data.get("id")
    if not raw_id:
        return None  # blank row

    # Build keyFeatures list from keyFeature_1 … keyFeature_N columns
    key_features = []
    for i in range(1, 10):
        kf = data.get(f"keyFeature_{i}")
        if kf:
            key_features.append(kf)

    # Build specs dict from spec_* columns
    specs = {}
    for k, v in data.items():
        if k.startswith("spec_") and v:
            specs[k[5:]] = v  # strip "spec_" prefix

    # Build boolFeatures dict from bool_* columns
    bool_features = {}
    for k, v in data.items():
        if k.startswith("bool_"):
            bv = to_bool(v)
            if bv is not None:
                bool_features[k[5:]] = bv  # strip "bool_" prefix

    # Core metadata booleans
    is_sponsored = to_bool(data.get("isSponsored")) or False
    is_green      = to_bool(data.get("isGreen"))     or False
    is_enabled    = to_bool(data.get("isEnabled"))
    if is_enabled is None:
        is_enabled = True  # default to enabled

    deal = {
        "id":              clean_id(raw_id),
        "providerName":    data.get("providerName ★") or data.get("providerName") or "Unknown",
        "planName":        data.get("planName ★")     or data.get("planName")     or "Standard Plan",
        "logoUrl":         data.get("logoUrl")         or "",
        "providerColor":   data.get("providerColor ★") or data.get("providerColor") or "0xFF000000",
        "description":     data.get("description")     or "",
        "keyFeatures":     key_features,
        "price":           to_float(data.get("price ★") or data.get("price")),
        "priceUnit":       data.get("priceUnit")        or "/mo",
        "affiliateUrl":    data.get("affiliateUrl")     or "",
        "directUrl":       data.get("directUrl")        or "",
        "rating":          to_float(data.get("rating"), 0.0),
        "isSponsored":     is_sponsored,
        "isGreen":         is_green,
        "isEnabled":       is_enabled,
        "applicableStates": split_states(data.get("applicableStates")),
        "tagline":         data.get("tagline")          or "",
        "tier":            data.get("tier")             or "",
        "specs":           specs,
        "boolFeatures":    bool_features,
        "details":         {},   # legacy field — left empty; populated by automation pipeline
    }

    return deal


def import_sheet(ws) -> list[dict]:
    """
    Read a category sheet and return a list of Deal dicts.
    Row 2 = headers, Row 3 = notes (skip), Row 4+ = data.
    """
    deals = []
    header_row = None

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        if row_idx == 2:
            # Extract headers from row 2
            header_row = [cell_val(c) or f"col_{i}" for i, c in enumerate(row)]
            continue
        if row_idx < 4:
            # Row 1 = title banner, Row 3 = notes — skip both
            continue
        if header_row is None:
            continue

        # Check if row is entirely empty
        values = [cell_val(c) for c in row]
        if all(v is None for v in values):
            continue

        row_dict = {header_row[i]: values[i] for i in range(min(len(header_row), len(values)))}
        deal = row_to_deal(row_dict)
        if deal:
            deals.append(deal)

    return deals


def merge_into_category(existing: list, incoming: list) -> tuple[list, int, int]:
    """
    Merge incoming deals into existing list by id.
    Returns (merged_list, added_count, updated_count).
    """
    existing_by_id = {d["id"]: d for d in existing}
    added = updated = 0

    for deal in incoming:
        if deal["id"] in existing_by_id:
            existing_by_id[deal["id"]] = deal
            updated += 1
        else:
            existing_by_id[deal["id"]] = deal
            added += 1

    # Preserve original order, then append new
    original_ids = [d["id"] for d in existing]
    merged = [existing_by_id[i] for i in original_ids if i in existing_by_id]
    for deal in incoming:
        if deal["id"] not in {d["id"] for d in merged}:
            merged.append(deal)

    return merged, added, updated


# ── Main ───────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Import SaveNest product template → products.json")
    parser.add_argument("--overwrite", action="store_true",
                        help="Fully replace category data (instead of merge)")
    parser.add_argument("--category", type=str, default=None,
                        help="Only import this category (e.g. electricity)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Preview output without writing to disk")
    args = parser.parse_args()

    if not TEMPLATE.exists():
        print(f"ERROR: Template not found at {TEMPLATE}")
        sys.exit(1)

    print(f"📂 Reading template: {TEMPLATE.name}")
    wb = openpyxl.load_workbook(TEMPLATE, data_only=True)

    # Load existing products.json
    if PRODUCTS_JSON.exists():
        with open(PRODUCTS_JSON) as f:
            products = json.load(f)
    else:
        products = {"metadata": {"lastUpdated": ""}}
        print("⚠️  products.json not found — creating from scratch")

    total_added = total_updated = 0

    for sheet_name, cat_key in SHEET_TO_CAT.items():
        if args.category and cat_key != args.category:
            continue
        if sheet_name not in wb.sheetnames:
            print(f"  ⏭  Sheet '{sheet_name}' not found — skipping")
            continue

        ws = wb[sheet_name]
        incoming = import_sheet(ws)

        if not incoming:
            print(f"  ⏭  {sheet_name}: no data rows found")
            continue

        if args.overwrite:
            products[cat_key] = incoming
            added, updated = len(incoming), 0
        else:
            existing = products.get(cat_key, [])
            merged, added, updated = merge_into_category(existing, incoming)
            products[cat_key] = merged

        total_added   += added
        total_updated += updated
        print(f"  ✓  {sheet_name} [{cat_key}]: {added} added, {updated} updated")

    print(f"\n📊 Summary: {total_added} added, {total_updated} updated")

    if args.dry_run:
        print("\n🔍 DRY RUN — first item of each updated category:")
        for sheet_name, cat_key in SHEET_TO_CAT.items():
            if cat_key in products and products[cat_key]:
                d = products[cat_key][0]
                print(f"\n  [{cat_key}] {d['id']} — {d['providerName']} / {d['planName']}")
                print(f"    price: {d['price']}  |  specs: {list(d.get('specs',{}).keys())[:4]}")
                print(f"    boolFeatures: {list(d.get('boolFeatures',{}).keys())[:4]}")
        print("\n✅ Dry run complete. No files written.")
        return

    # Write updated products.json
    with open(PRODUCTS_JSON, "w") as f:
        json.dump(products, f, indent=2, ensure_ascii=False)
    print(f"\n✅ Written → {PRODUCTS_JSON}")


if __name__ == "__main__":
    main()
